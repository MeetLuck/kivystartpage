import os,csv, shutil
from socket import gethostname
from pathlib import Path
from datetime import datetime,timedelta
from functools import lru_cache
from time import time
from openpyxl import load_workbook
from itertools import product
from shutil import copy2

def timeit(func):
    from functools import wraps
    @wraps(func)
    def timeit_wrapper(*args, **kwargs):
        import time
        start_time = time.perf_counter()
        result = func(*args, **kwargs)
        end_time = time.perf_counter()
        total_time = end_time - start_time
        print(f'#Function {func.__name__} took {total_time:.4f} seconds')
        return result
    return timeit_wrapper

if gethostname() == 'main':
    # NOTE files
    minwonfile = r'D:\01.업무문서\000. 민원처리내역.xlsx'
    wonhyungfile = r'D:\01.업무문서\00.원형도로 단속일지(최종본).xlsx'
    todofile =  r'D:\01.A조\sub박종우\banpo\todo.txt'
    minwonpyfile = r'\\Sub\d\01.A조\sub박종우\banpo\minwon\minwon.py'
    minwondbfile =  r'D:\A조\박종우\banpo\minwon\minwondb.py'
    dbbrowserfile = r'D:\A조\박종우\DB Browser for SQLite\DB Browser for SQLite.exe'
    hpscanfile =  r'C:\Program Files (x86)\HP\HP OfficeJet Pro 8710\bin\HPScan.exe'
    snippingtoolfile =  r'C:\WINDOWS\system32\SnippingTool.exe'
    gitfile = r'\\Sub\d\A조\sub박종우\GitPortable\GitBashPortable.exe'
    #NOTE png files
    #daypatrolfile =  r'\\Sub\d\01.A조\sub박종우\banpo\day_patrol.png'
    #nightpatrolfile =  r'\\Sub\d\01.A조\sub박종우\banpo\night_patrol.png'
    #dongpostfile =  r'\\Sub\d\01.A조\sub박종우\banpo\images\dong_post.png'
    # NOTE folders
    workfolder =  r'D:\01.업무문서'
    commutefolder =  r'D:\01.업무문서\04.출근부 관련\02.각 조별 월별 출근부'
    weeklyworkfolder =  r'D:\01.업무문서\05.주간업무보고서 관련\2022년\주간'
    incidentfolder =  r'D:\01.업무문서\03.사건사고 보고서'
    Afolder =  r'D:\A조'
    excellinkfolder =  r'D:\A조\엑셀링크'
    parkfolder =  r'D:\A조\박종우'
    cctvfolder =  r'D:\A조\박종우\CCTV'
    infofolder = r'D:\A조\안내문'

    #NOTE png files
    #dongpostfile    = r'\\Sub\d\01.A조\sub박종우\banpo\dong_post.png'
    dongpostfile    = r'D:\A조\박종우\banpo\dong_post.png'
    phonefile       = r'D:\A조\직원연락처(내선번호)-3.11.jpg'
    map1ffile       = r'D:\A조\엑셀링크\반포자이아파트.jpg'
    mapb1file       = r'D:\A조\엑셀링크\지하주차장.png'
    mapb1cctvfile   = r'D:\A조\엑셀링크\지하주차장cctv.png'
    vaultfile       = r'D:\A조\엑셀링크\금고.png'
    gatemanfile     = r'D:\A조\박종우\subbanpo\세대현관문.jpg'

    #NOTE image_filename
    ev_button_img       = r'\\Sub\d\01.A조\sub박종우\banpo\ev_button.png'
    cctv_button_img     = r'\\Sub\d\01.A조\sub박종우\banpo\cctv_button.png'
    homenet_button_img  = r'\\Sub\d\01.A조\sub박종우\banpo\homenet_button.png'
    vault_button_img    = r'\\Sub\d\01.A조\sub박종우\banpo\vault_button.png'
elif gethostname() == 'Sub':
    # NOTE files
    minwonfile  = r'\\Main\d\\01.업무문서\000. 민원처리내역.xlsx'
    wonhyungfile = r'\\Main\d\01.업무문서\00.원형도로 단속일지(최종본).xlsx'
    todofile = r'D:\01.A조\sub박종우\banpo\todo.txt'
    minwonpyfile = r'D:\01.A조\sub박종우\banpo\minwon\minwon.py'
    minwondbfile = r'D:\01.A조\sub박종우\banpo\minwon\minwondb.py'
    hpscanfile = r'C:\Program Files (x86)\HP\HP OfficeJet Pro 8710\bin\HPScan.exe'
    snippingtoolfile = r'C:\WINDOWS\system32\SnippingTool.exe'
    dbbrowserfile = r'D:\01.A조\sub박종우\DB Browser for SQLite\DB Browser for SQLite.exe'
    gitfile = r'D:\01.A조\sub박종우\GitPortable\GitBashPortable.exe'

    # NOTE folders
    workfolder =  r'\\Main\d\01.업무문서'
    commutefolder =  r'\\Main\d\01.업무문서\04.출근부 관련\02.각 조별 월별 출근부'
    weeklyworkfolder =  r'\\Main\d\01.업무문서\05.주간업무보고서 관련\2022년\주간'
    incidentfolder =  r'\\Main\d\01.업무문서\03.사건사고 보고서'
    Afolder =  r'\\Main\d\A조'
    excellinkfolder =  r'\\Main\d\A조\엑셀링크'
    parkfolder =  r'\\Main\d\A조\박종우'
    cctvfolder = r'\\Main\d\A조\박종우\CCTV'
    infofolder =  r'\\Main\d\A조\안내문'
    banpofolder = r'D:\01.A조\sub박종우\banpo'

    #NOTE banpo\images png files
    dongpostfile    = r'D:\01.A조\sub박종우\banpo\images\dong_post.png'
    phonefile       = r'\\Main\d\A조\직원연락처(내선번호),조직도\직원연락처(내선번호)-5-02.jpg'
    map1ffile       = r'\\Main\d\A조\엑셀링크\반포자이아파트.jpg'
    mapb1file       = r'\\Main\d\A조\엑셀링크\\지하주차장.png'
    mapb1cctvfile   = r'\\Main\d\A조\엑셀링크\\지하주차장cctv.png'
    vaultfile       = r'\\Main\d\A조\엑셀링크\\금고.png'
    gatemanfile     = r'D:\01.A조\sub박종우\banpo\images\세대현관문.jpg'

    #NOTE Button image_filename
    ev_button_img       = r'D:\01.A조\sub박종우\banpo\images\ev_button.png'
    cctv_button_img     = r'D:\01.A조\sub박종우\banpo\images\cctv_button.png'
    homenet_button_img  = r'D:\01.A조\sub박종우\banpo\images\homenet_button.png'
    vault_button_img    = r'D:\01.A조\sub박종우\banpo\images\vault_button.png'
elif gethostname() == 'meetluck':
    # NOTE files
    minwonfile = r'D:\Documents\d\01.업무문서\000. 민원처리내역.xlsx'
    wonhyungfile = r'\\Main\d\01.업무문서\00.원형도로 단속일지(최종본).xlsx'
    todofile = r'D:\Documents\banpo\todo.txt'
    minwodbfile =  r'D:\01.A조\sub박종우\banpo\minwon\minwondb.py'
    dbbrowserfile = r'D:\01.A조\sub박종우\DB Browser for SQLite\DB Browser for SQLite.exe'
    hpscanfile =  r'C:\Program Files (x86)\HP\HP OfficeJet Pro 8710\bin\HPScan.exe'
    snippingtoolfile =  r'C:\WINDOWS\system32\SnippingTool.exe'
    gitfile = r'D:\01.A조\sub박종우\GitPortable\GitBashPortable.exe'

    #NOTE png files
    dongpostfile =  r'D:\Documents\banpo\images\dong_post.png'

    # NOTE folders
    workfolder =  r'D:\Documents\d\01.업무문서'
    commutefolder =  r'D:\Documents\d\01.업무문서\04.출근부 관련\02.각 조별 월별 출근부'
    weeklyworkfolder =  r'D:\Documents\d\01.업무문서\05.주간업무보고서 관련\2022년\주간'
    incidentfolder =  r'D:\Documents\d\01.업무문서\03.사건사고 보고서'
    Afolder =  r'D:\Documents\d\A조'
    excellinkfolder =  r'D:\Documents\d\A조\엑셀링크'
    parkfolder =  r'D:\Documents\d\A조\박종우'
    cctvfolder =  r'D:\Documents\d\A조\박종우\CCTV'
    infofolder =  r'D:\Documents\d\A조\안내문'

    #NOTE image_filename
    ev_button_img       = r'D:\Documents\banpo\ev_button.png'
    cctv_button_img     = r'D:\Documents\banpo\cctv_button.png'
    homenet_button_img  = r'D:\Documents\banpo\homenet_button.png'
    vault_button_img    = r'D:\Documents\banpo\vault_button.png'
else:
    raise Exception(r'unknown host {}'.format(gethostname()))

# NOTE get day1 for A조
A_day1 = datetime(2022,1,31)

today = datetime.today()
default_date=(today.month,today.day, today.year)

print( " ==> import starthelp.py ")

def getminwonfile(root):
    copiedfile = Path(root,'files/minwon2022.xlsx')
    print(copiedfile)
    if file_exists(copiedfile):
        if is_modified(minwonfile,copiedfile):
            shutil.copy2(minwonfile,copiedfile)
            print(f'{minwonfile} modified -> {copiedfile} copied')
    else:
        shutil.copy2(minwonfile,copiedfile)
        print(f'{copiedfile} not exists, {copiedfile} copied')
    return copiedfile

def file_exists(file):
    from os.path import exists
    return exists(file)

def is_modified(origin,copied):
    return os.stat(origin).st_mtime != os.stat(copied).st_mtime
def get_weekday(date):
    days = '월 화 수 목 금 토 일'.split()
    return days[date.weekday()]
def get_weeklyreports(date,weekday=3):
    from calendar import monthrange
    weeklydates = []
    firstweekday,end = monthrange(date.year, date.month)
    for day in range(1,end+1):
        adate = datetime(date.year, date.month, day)
        if adate.weekday() == weekday:  # 3 -> 목, 4->금
            weeklydates.append(adate)
    return weeklydates

@lru_cache(maxsize=4)
def get_filenamesOndate(date):
    yyyymm = date.strftime("%Y%m")  #202205
    month  = date.strftime("%m")    #05(zero-padding)
    date1  = date.strftime("%Y%m%d")
    date2  = (date + timedelta(days=1)).strftime("%Y%m%d")
    #print('==>',gethostname())
    if gethostname() == 'main': 
        monthlycommutefile = r'D:\01.업무문서\04.출근부 관련\02.각 조별 월별 출근부\{} A,B,C조별 월별출근부(반포자이).xlsx'.format(yyyymm)
        dailyreportfile1 = r'D:\01.업무문서\01.일일상황보고(상황실)\2022년{}월\일일 상황 보고 {}.xlsx'.format(month,date1)
        dailyreportfile2 = r'D:\01.업무문서\01.일일상황보고(상황실)\2022년{}월\일일 상황 보고 {}.xlsx'.format(month,date2)
    elif gethostname() == 'Sub': 
        monthlycommutefile = r'\\Main\d\01.업무문서\04.출근부 관련\02.각 조별 월별 출근부\{} A,B,C조별 월별출근부(반포자이).xlsx'.format(yyyymm)
        dailyreportfile1 = r'\\Main\d\01.업무문서\01.일일상황보고(상황실)\2022년{}월\일일 상황 보고 {}.xlsx'.format(month,date1)
        dailyreportfile2 = r'\\Main\d\01.업무문서\01.일일상황보고(상황실)\2022년{}월\일일 상황 보고 {}.xlsx'.format(month,date2)
    elif gethostname() == 'meetluck': 
        monthlycommutefile = r'D:\Documents\d\01.업무문서\04.출근부 관련\02.각 조별 월별 출근부\{} A,B,C조별 월별출근부(반포자이).xlsx'.format(yyyymm)
        dailyreportfile1 = r'D:\Documents\d\01.업무문서\01.일일상황보고(상황실)\2022년{}월\일일 상황 보고 {}.xlsx'.format(month,date1)
        dailyreportfile2 = r'D:\Documents\d\01.업무문서\01.일일상황보고(상황실)\2022년{}월\일일 상황 보고 {}.xlsx'.format(month,date2)
    else:
        raise Exception(f'invalid hostname: {gethostname()}')
    from pathlib import Path
    if not Path(monthlycommutefile).is_file():
        raise Exception(f'{__file__}\n  {monthlycommutefile} not exist')
    return (monthlycommutefile,dailyreportfile1,dailyreportfile2)

def get_monthlycommutefile(date):
    return get_filenamesOndate(date)[0]
def get_dailyreportfile(date):    return get_filenamesOndate(date)[1:3]

@timeit
def create_commute_csv(date):
    #XXX openpyxl in read_only mode , wb.close not work
    month,day = date.month, date.day
    commutefile = get_monthlycommutefile(date)
    copiedfile = r'./commute{}.xlsx'.format(date.month)
    csvfile = r'./commute{}.csv'.format(date.month)
    if file_exists(copiedfile) and not is_modified(commutefile,copiedfile):
        print(f'{copiedfile} exist and not modified, check csv file')
        if file_exists(csvfile):
            print(f'{csvfile} exist, nothing changed')
            return

    print(f'creating {csvfile}')
    copy2(commutefile,copiedfile)
    start = time()
    wb = load_workbook( copiedfile, data_only=True,read_only=True )
    ws = wb['조별연차표'] 

    def R1C1(row,col): return ws.cell(row,col).value
    # XXX to reduce elapsed time, get team and off for date1,date2,date3,date4 at once
    #  30     A1, 이종화,     B1, 서광석
    #  31     A2 이종화,      B2, 류정희
    #   1     A2 이종화,      B2, 류정희
    #   2     A2 이종화,      B2, 류정희

    #class team: day,night = get_workteam(date)
    #class off: day,night = '',''
    with open(csvfile, 'w',newline='') as f:
        write = csv.writer(f)
        #write.writerows( get_rows() )
        for r,c in product([4,8,12,16,20],[2,3,4,5,6,7,8]):
            offlist = []
            aday = R1C1(r,c)
            if not aday: continue
            print(aday)
            adate = datetime(date.year,date.month,aday)
            offlist.append(aday)                   # date
            offlist.append(get_workteam(adate)[0]) # team for day
            offlist.append(R1C1(r+1,c))            # off  for day
            offlist.append(get_workteam(adate)[1]) # team for night
            offlist.append(R1C1(r+2,c))            # off  for night
            write.writerow( offlist )

@lru_cache(maxsize=4)
def get_offworkers(date):
    create_commute_csv(date)
    csvfile = f'commute{date.month}.csv'
    with open(csvfile, 'r') as file:
        reader = csv.reader(file)
        for r1,row in enumerate( row for row in reader ):
            if int(date.day) == int(r1)+1:
                return row[:]
    # XXX to reduce elapsed time, get team and off for date1,date2,date3,date4 at once
    #  1     A1, 이종화,     B1, 서광석
    #  2     A2 이종화,      B2, 류정희

def convert_offworkers(date:datetime) -> list: # [6/1, A1, 이종화 B1 육백근]
    offs = list()
    for i in range(4):
        adate = date + timedelta(days=i)
        aday = f'{adate.month}/{adate.day}'
        offs.append([aday] + get_offworkers(adate)[1:] )
    return offs

def get_dayornight(date):
    diff = (datetime(date.year,date.month,date.day) - A_day1).days
    dayornight=['day1','day2','night1','night2','off1','off2']
    return dayornight[diff % 6]
def get_workteam(date):
    diff = (datetime(date.year,date.month,date.day) - A_day1).days
    teams = 'A1,B1 A2,B2 C1,A1 C2,A2 B1,C1 B2,C2'.split()
    dayteam,nightteam = teams[diff % 6].split(',')
    return (dayteam,nightteam)

if __name__ == '__main__':
    date = datetime.strptime('22-5-10','%y-%m-%d')
    create_commute_csv(date)
    workteam = get_offworkers(date)
    print(today,workteam)
